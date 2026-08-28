from __future__ import annotations

from datetime import timedelta

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from tender_ai.export import export_search_session
from tender_ai.extractors.runner import DOCUMENT_PARSER_VERSION, EXTRACTION_VERSION, ExtractionRunner
from tender_ai.review import ensure_review_item
from tender_ai.search import SearchRequest, build_source_plan
from tender_ai.snapshots.store import SnapshotStore
from tender_ai.status.time import now_shanghai
from tender_ai.storage.database import create_engine_for, initialize_database, refresh_tender_fts, search_projects, session_scope
from tender_ai.storage.models import Announcement, CodexReviewItem, DocumentParse, Project, SearchSession, SearchSessionProject
from tender_ai.templates import list_templates, save_template
from tender_ai.urls import content_hash
from tender_ai.web.app import create_app


def test_structured_request_round_trip_and_dynamic_source_plan():
    request = SearchRequest(region="内蒙古自治区", regions=("内蒙古自治区",), industries=("solar", "storage"), days=30)
    restored = SearchRequest.from_dict(request.to_dict())
    assert restored.region == "内蒙古自治区"
    assert restored.industries == ("solar", "storage")
    plan = build_source_plan(request)
    selected = {row["source_id"] for row in plan if row["selected"]}
    assert selected == {"ccgp", "ggzy", "cebpubservice"}
    assert all(row["region"] == "全国" for row in plan if row["selected"])


def test_template_and_open_excel_export(tmp_path):
    database = tmp_path / "stage5.db"
    engine = initialize_database(create_engine_for(database))
    with session_scope(engine) as session:
        project = Project(
            project_id="stage5-project", project_name="测试储能项目", status="OPEN", status_reason="OPEN_DOCUMENT_DOWNLOAD_ACTIVE",
            province="内蒙古自治区", city="呼和浩特市", owner="测试招标人", source_level="A",
            bid_deadline=now_shanghai() + timedelta(days=3), updated_at=now_shanghai(), created_at=now_shanghai(),
        )
        session.add(project)
        session.flush()
        announcement = Announcement(project_id=project.project_id, title="测试公告", source_url="https://example.test/a")
        session.add(announcement)
        session.flush()
        session.add(SearchSession(session_id="stage5-session", request_json="{}", status="COMPLETED", open_count=1, candidate_count=1))
        session.add(SearchSessionProject(session_id="stage5-session", project_id=project.project_id, announcement_id=announcement.id, status_at_search="OPEN"))
        refresh_tender_fts(session, project, "测试储能项目")
    saved = save_template("测试模板", SearchRequest(region="内蒙古自治区", industries=("storage",)), database=str(database))
    assert saved["request"]["industries"] == ["storage"]
    assert list_templates(database=str(database))[0]["name"] == "测试模板"
    output = export_search_session("stage5-session", database=str(database), output_path=tmp_path / "result.xlsx")
    assert output["count"] == 1
    workbook = load_workbook(output["path"])
    assert workbook.active.max_row == 2
    assert workbook.active.cell(1, 1).value == "省"


def test_web_data_browser_routes_do_not_search(tmp_path):
    database = tmp_path / "web.db"
    engine = initialize_database(create_engine_for(database))
    with session_scope(engine) as session:
        session.add(Project(project_id="web-project", project_name="Web 测试项目", status="UNKNOWN", status_reason="UNKNOWN_NO_PARTICIPATION_DEADLINE", created_at=now_shanghai(), updated_at=now_shanghai()))
    client = TestClient(create_app(database=str(database)))
    assert client.get("/healthz").json()["scheduling_mode"] == "ON_DEMAND_ONLY"
    assert client.get("/").status_code == 200
    assert client.get("/projects").status_code == 200
    assert client.get("/projects/web-project").status_code == 200
    assert client.get("/settings").status_code == 200
    assert client.get("/sources").status_code == 200
    assert client.get("/review").status_code == 200
    assert client.get("/api/projects?limit=1").json()["count"] == 1
    assert client.get("/api/projects/web-project").status_code == 200


def test_search_reuses_unchanged_document_extraction_cache(tmp_path):
    database = tmp_path / "cache.db"
    engine = initialize_database(create_engine_for(database))
    with session_scope(engine) as session:
        project = Project(
            project_id="cache-project",
            project_name="缓存储能项目",
            status="UNKNOWN",
            status_reason="UNKNOWN_NO_PARTICIPATION_DEADLINE",
            province="内蒙古自治区",
            extraction_version=EXTRACTION_VERSION,
            created_at=now_shanghai(),
            updated_at=now_shanghai(),
        )
        session.add(project)
        session.flush()
        announcement = Announcement(
            project_id=project.project_id,
            title="缓存储能项目公告",
            source_url="https://example.test/cache",
            content_hash=content_hash("缓存储能项目公告"),
            extraction_status="SUCCESS",
            extraction_version=EXTRACTION_VERSION,
        )
        session.add(announcement)
        session.flush()
        snapshot = SnapshotStore(tmp_path / "snapshots").save_text(
            session,
            source_url=announcement.source_url,
            text="缓存储能项目公告",
            content_type="text/html",
            announcement_id=announcement.id,
        )
        announcement.snapshot_id = snapshot.snapshot_id
        session.add(
            DocumentParse(
                announcement_id=announcement.id,
                project_id=project.project_id,
                content_hash=snapshot.sha256,
                content_type="text/html",
                parser="html.beautifulsoup_rule",
                parser_version=DOCUMENT_PARSER_VERSION,
                parse_status="SUCCESS",
                text_length=8,
            )
        )
    summary = ExtractionRunner(database=str(database)).run(sample_size=1, consolidate=False, reuse_cached=True)
    assert summary.extraction_cache_hits == 1
    assert summary.pdf_count == 0


def test_chinese_project_search_falls_back_when_fts_tokenizer_misses(tmp_path):
    engine = initialize_database(create_engine_for(tmp_path / "fts.db"))
    with session_scope(engine) as session:
        project = Project(
            project_id="fts-project",
            project_name="哈密储能系统采购",
            status="UNKNOWN",
            created_at=now_shanghai(),
            updated_at=now_shanghai(),
        )
        session.add(project)
        session.flush()
        refresh_tender_fts(session, project, project.project_name)
        assert search_projects(session, "哈密 储能", limit=10) == [project.project_id]


def test_existing_review_is_written_after_session_reassociation(tmp_path):
    database = tmp_path / "review-output.db"
    engine = initialize_database(create_engine_for(database))
    with session_scope(engine) as session:
        project = Project(
            project_id="review-output-project",
            project_name="待确认项目",
            status="UNKNOWN",
            status_reason="UNKNOWN_NO_PARTICIPATION_DEADLINE",
            created_at=now_shanghai(),
            updated_at=now_shanghai(),
        )
        session.add(project)
        session.flush()
        announcement = Announcement(project_id=project.project_id, title="待确认项目公告", source_url="https://example.test/review-output")
        session.add(announcement)
        session.flush()
        session.add(SearchSession(session_id="review-output-session", request_json="{}", status="RUNNING"))
        ensure_review_item(session, project, announcement=announcement)
        ensure_review_item(session, project, announcement=announcement, search_session_id="review-output-session")
        session.flush()
        item = session.scalar(select(CodexReviewItem).where(CodexReviewItem.search_session_id == "review-output-session"))
        assert item is not None
