"""SearchSession 结果导出为轻量 Excel。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from tender_ai.config_loader import APP_ROOT
from tender_ai.status.time import as_shanghai, now_shanghai
from tender_ai.storage.database import create_engine_for, initialize_database, session_scope
from tender_ai.storage.models import Announcement, Project, SearchSession, SearchSessionProject


EXPORT_HEADERS = [
    ("省", "province"),
    ("市", "city"),
    ("县/旗", "county"),
    ("项目名称", "project_name"),
    ("招标人", "owner"),
    ("项目类型", "project_type"),
    ("规模", "scale"),
    ("预算", "budget"),
    ("报名截止", "registration_deadline"),
    ("文件截止", "document_deadline"),
    ("投标截止", "bid_deadline"),
    ("开标", "open_time"),
    ("剩余小时", "remaining_hours"),
    ("状态", "status"),
    ("状态原因", "status_reason"),
    ("来源等级", "source_level"),
    ("来源", "source_name"),
    ("原始链接", "source_url"),
]


def _display(value: Any) -> Any:
    if isinstance(value, datetime):
        return as_shanghai(value).strftime("%Y-%m-%d %H:%M:%S")
    return value


def _remaining_hours(project: Project) -> float | None:
    deadlines = [
        getattr(project, field_name, None)
        for field_name in ("qualification_deadline", "registration_deadline", "document_deadline", "bid_deadline", "open_time")
    ]
    dates = [as_shanghai(value) for value in deadlines if isinstance(value, datetime)]
    if not dates:
        return None
    return round((min(dates) - now_shanghai()).total_seconds() / 3600, 2)


def export_search_session(session_id: str, *, database: str | None = None, include_unknown: bool = False, output_path: str | Path | None = None) -> dict[str, Any]:
    """导出一次已完成搜索；默认只导出 OPEN，避免把历史 CLOSED 混给用户。"""
    engine = initialize_database(create_engine_for(database))
    with session_scope(engine) as session:
        search_session = session.get(SearchSession, session_id)
        if search_session is None:
            raise KeyError(f"Search Session 不存在: {session_id}")
        allowed = {"OPEN", "UNKNOWN"} if include_unknown else {"OPEN"}
        rows = list(
            session.scalars(
                select(SearchSessionProject)
                .where(SearchSessionProject.session_id == session_id, SearchSessionProject.status_at_search.in_(allowed))
                .order_by(SearchSessionProject.status_at_search, SearchSessionProject.id)
            ).all()
        )
        projects: list[tuple[Project, Announcement | None]] = []
        for link in rows:
            project = session.get(Project, link.project_id)
            if project is None or project.ignored:
                continue
            announcement = session.get(Announcement, link.announcement_id) if link.announcement_id else None
            projects.append((project, announcement))
    path = Path(output_path) if output_path else APP_ROOT.parent / "output" / f"search_{session_id}.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "搜索结果"
    header_fill = PatternFill("solid", fgColor="173F4D")
    header_font = Font(color="FFFFFF", bold=True)
    for column, (label, _) in enumerate(EXPORT_HEADERS, 1):
        cell = sheet.cell(1, column, label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_index, (project, announcement) in enumerate(projects, 2):
        values = {
            "province": project.province,
            "city": project.city,
            "county": project.county,
            "project_name": project.project_name,
            "owner": project.owner,
            "project_type": project.project_type,
            "scale": project.project_scale or project.capacity_mw or project.capacity_mwh,
            "budget": str(project.budget) if project.budget is not None else None,
            "registration_deadline": project.registration_deadline,
            "document_deadline": project.document_deadline,
            "bid_deadline": project.bid_deadline,
            "open_time": project.open_time,
            "remaining_hours": _remaining_hours(project),
            "status": next((link.status_at_search for link in rows if link.project_id == project.project_id), project.status),
            "status_reason": project.status_reason,
            "source_level": project.source_level,
            "source_name": project.source_name,
            "source_url": (announcement.source_url if announcement else None) or project.source_url,
        }
        for column, (_, key) in enumerate(EXPORT_HEADERS, 1):
            sheet.cell(row_index, column, _display(values.get(key)))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 26
    for index, (label, _) in enumerate(EXPORT_HEADERS, 1):
        width = max(12, min(42, len(label) + 4, max((len(str(sheet.cell(row, index).value or "")) for row in range(1, min(sheet.max_row, 20) + 1)), default=12) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(path)
    return {"session_id": session_id, "path": str(path), "count": len(projects), "include_unknown": include_unknown}


__all__ = ["export_search_session"]
